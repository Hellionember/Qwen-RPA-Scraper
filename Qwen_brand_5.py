import os
import re
from dotenv import load_dotenv
import pandas as pd
import sqlite3
import logging
import time
import random
import pyperclip
from pathlib import Path
from tqdm import tqdm
from datetime import datetime
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException


# Загружаем переменные окружения из файла .env
load_dotenv()

# ==================== ЗАГРУЗКА КОНФИГУРАЦИИ ====================
config = {}
config_file = 'config.txt'

try:
    with open(config_file, 'r', encoding='utf-8') as f:
        for line in f:
            if '=' in line and not line.startswith('#'):
                key, val = line.strip().split('=', 1)
                config[key.strip()] = val.strip()
except FileNotFoundError:
    print(f"[WARNING] {config_file} не найден, используются значения по умолчанию.")

# Настройка логирования
logging.basicConfig(
    level=getattr(logging, config.get('LOG_LEVEL', 'INFO').upper(), logging.INFO),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(config.get('LOG_FILE', 'processing.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Отключение избыточного DEBUG-логирования от библиотек Selenium и urllib3 в консоли
for logger_name in ["urllib3", "selenium", "urllib3.connectionpool", "selenium.webdriver.remote.remote_connection", "http.client"]:
    logging.getLogger(logger_name).setLevel(logging.WARNING)


# Безопасная загрузка учетных данных
QWEN_EMAIL = os.getenv('QWEN_EMAIL')
QWEN_PASSWORD = os.getenv('QWEN_PASSWORD')

if not QWEN_EMAIL or not QWEN_PASSWORD:
    logger.error("❌ Учетные данные не найдены! Проверьте наличие файла .env")
    exit(1)

# ==================== ПАРАМЕТРЫ ====================
INPUT_EXCEL = config.get('INPUT_EXCEL', 'Входной.xlsx')
DB_PATH = config.get('DB_PATH', 'materials.db')
OUTPUT_EXCEL = config.get('OUTPUT_EXCEL', 'Выходной.xlsx')
OUTPUT_MISSING_BRAND = config.get('OUTPUT_MISSING_BRAND', 'Выходной_бренда_нет_в_справочнике.xlsx')
EDGE_DRIVER_PATH = config.get('EDGE_DRIVER_PATH', r"C:\drivers\msedgedriver.exe")
LOG_FILE = config.get('LOG_FILE', 'processing.log')
DEBUG_DIR = config.get('DEBUG_DIR', 'debug')

BATCH_SIZE = int(config.get('BATCH_SIZE', '100'))
DELAY_MIN = int(config.get('DELAY_MIN', '5'))
DELAY_MAX = int(config.get('DELAY_MAX', '15'))
NEW_CHAT_EVERY_N_BATCHES = int(config.get('NEW_CHAT_EVERY_N_BATCHES', '10'))
MAX_RETRIES = int(config.get('MAX_RETRIES', '3'))          
MAX_TIMEOUT_RETRIES = int(config.get('MAX_TIMEOUT_RETRIES', '3'))
RESPONSE_TIMEOUT = int(config.get('RESPONSE_TIMEOUT', '600'))

WAIT_FOR_ELEMENT_TIMEOUT = int(config.get('WAIT_FOR_ELEMENT_TIMEOUT', '15'))
WAIT_FOR_NEW_CHAT_TIMEOUT = int(config.get('WAIT_FOR_NEW_CHAT_TIMEOUT', '15'))
SWITCH_MODE_TIMEOUT = int(config.get('SWITCH_MODE_TIMEOUT', '15'))
LOGIN_TIMEOUT = int(config.get('LOGIN_TIMEOUT', '15'))

AFTER_CLICK_DELAY = int(config.get('AFTER_CLICK_DELAY', '3'))
AFTER_CLEAR_DELAY = int(config.get('AFTER_CLEAR_DELAY', '2'))
AFTER_PASTE_DELAY = int(config.get('AFTER_PASTE_DELAY', '3'))
AFTER_ENTER_DELAY = int(config.get('AFTER_ENTER_DELAY', '8'))
EXTRA_RESPONSE_DELAY = int(config.get('EXTRA_RESPONSE_DELAY', '8'))
PAGE_LOAD_DELAY = int(config.get('PAGE_LOAD_DELAY', '15'))

SAVE_DEBUG_FILES = config.get('SAVE_DEBUG_FILES', 'True').lower() == 'true'
AUTO_CREATE_NEW_CHAT = config.get('AUTO_CREATE_NEW_CHAT', 'True').lower() == 'true'
HEADLESS_MODE = config.get('HEADLESS_MODE', 'False').lower() == 'true'
WINDOW_WIDTH = int(config.get('WINDOW_WIDTH', '1920'))
WINDOW_HEIGHT = int(config.get('WINDOW_HEIGHT', '1080'))

CSV_SEPARATOR = config.get('CSV_SEPARATOR', '€')
EXPECTED_ENCODING = config.get('EXPECTED_ENCODING', 'utf-8')

RETRY_BATCH_SIZE = int(config.get('RETRY_BATCH_SIZE', '50'))
MAX_RETRY_PER_ROW = int(config.get('MAX_RETRY_PER_ROW', '3'))

SESSION_START = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
_session_logged_for_class = set()

# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================
def clean_id(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        return val_str[:-2]
    return val_str

def safe_save_workbook(book, file_path):
    """Безопасно сохраняет рабочую книгу Excel. Если файл заблокирован,
    скрипт ждет его закрытия пользователем вместо завершения с ошибкой."""
    while True:
        try:
            book.save(file_path)
            break
        except PermissionError:
            logger.warning(
                f"⚠️ ОШИБКА ДОСТУПА: Файл '{file_path}' открыт в Excel или заблокирован другой программой! "
                "Пожалуйста, закройте этот файл, чтобы скрипт мог записать новые результаты. "
                "Ожидание закрытия и повторная попытка через 10 секунд..."
            )
            time.sleep(10)
        except Exception as e:
            logger.error(f"❌ Неожиданная ошибка сохранения книги '{file_path}': {e}")
            raise e

def load_already_processed_from_excel(excel_path, db_path):
    """Считывает уже готовые ид_исторический из выходного файла и заносит их в БД как обработанные"""
    if not Path(excel_path).exists():
        return
    try:
        wb = load_workbook(excel_path, read_only=True)
        processed_ids = set()
        ws = wb.active
        if ws is None:
            return
        
        id_col_indices = []
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            if 'ид_исторический' in row:
                id_col_indices = [i for i, val in enumerate(row) if val == 'ид_исторический']
                continue
            for idx in id_col_indices:
                if idx < len(row) and row[idx] is not None:
                    val_clean = clean_id(row[idx])
                    if val_clean and val_clean != 'ид_исторический':
                        processed_ids.add(val_clean)
        
        if processed_ids:
            logger.info(f"🔎 Найдено {len(processed_ids)} ранее обработанных записей в {excel_path}.")
            with sqlite3.connect(db_path) as conn:
                processed_ids_list = list(processed_ids)
                chunk_size = 500
                matched_numbers = []
                for i in range(0, len(processed_ids_list), chunk_size):
                    chunk = processed_ids_list[i:i+chunk_size]
                    placeholders = ','.join('?' for _ in chunk)
                    query = f"SELECT номер, класс, бренд FROM materials WHERE CAST(ид_исторический AS TEXT) IN ({placeholders})"
                    rows = conn.execute(query, chunk).fetchall()
                    matched_numbers.extend(rows)
                
                if matched_numbers:
                    conn.executemany(
                        "INSERT OR IGNORE INTO processed (номер, class_name, brand) VALUES (?, ?, ?)",
                        [(row[0], row[1], row[2]) for row in matched_numbers]
                    )
                    logger.info(f"⏭️ Успешно пропущено {len(matched_numbers)} записей, так как они уже есть в файле результатов.")
    except Exception as e:
        logger.warning(f"⚠️ Ошибка при импорте уже обработанных данных из Excel: {e}")

def has_error_message(text):
    """Проверяет ответ на наличие сообщений об ошибках в интерфейсе модели"""
    if not text:
        return False
    error_keywords = [
        "что-то пошло не так",
        "something went wrong",
        "пожалуйста, попробуйте еще раз",
        "please try again",
        "ошибка генерации",
        "failed to generate",
        "сервис временно недоступен",
        "an error occurred",
        "try again later",
        "server busy",
        "превышен лимит",
        "rate limit",
        # Системные ошибки облака Qwen / Alibaba Cloud
        "allocated quota exceeded",
        "issue connecting to",
        "quota limit",
        "oops!"
    ]
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in error_keywords)

def ensure_driver_active(driver_container):
    """Проверяет работоспособность сессии браузера. Если сессия неактивна
    (браузер упал или закрыт), полностью пересоздает его."""
    driver = driver_container[0]
    session_active = False
    try:
        # Простой запрос для проверки, откликается ли драйвер
        _ = driver.title
        session_active = True
    except Exception:
        session_active = False
    
    if not session_active:
        logger.warning("⚠️ Сессия браузера неактивна или повреждена. Выполняется автоматический перезапуск браузера и повторная авторизация...")
        try:
            driver.quit()
        except:
            pass
        new_driver = setup_webdriver()
        if new_driver:
            driver_container[0] = new_driver
            new_driver.get("https://chat.qwen.ai")
            time.sleep(PAGE_LOAD_DELAY)
            if login_to_qwen(new_driver):
                switch_to_fast_mode(new_driver)
                return True
        return False
    return True

def handle_dual_choices(driver):
    """Обнаруживает кнопки двойного выбора (I prefer this response / smulti-make-better) в интерфейсе и случайно кликает на одну"""
    try:
        # 1. Поиск кнопок выбора по быстрому CSS-селектору класса smulti-make-better
        option_buttons = driver.find_elements(By.CSS_SELECTOR, "button.smulti-make-better")
        
        # Оставляем только видимые и доступные для взаимодействия кнопки
        option_buttons = [btn for btn in option_buttons if btn.is_displayed() and btn.is_enabled()]
        
        if len(option_buttons) >= 2:
            chosen_btn = random.choice(option_buttons[:2])
            chosen_text = chosen_btn.text.strip() or "I prefer this response"
            logger.info(f"⚖️ Обнаружено сравнение вариантов (Qwen Studio feedback). Выбираем случайный ответ: '{chosen_text}'")
            
            # JS-клик обходит возможные перекрытия слоев в интерфейсе
            driver.execute_script("arguments[0].click();", chosen_btn)
            time.sleep(AFTER_CLICK_DELAY)
            return True

        # 2. Фолбек: если глобальные кнопки сравнения по CSS-классу не найдены, проверяем стандартные кнопки вариантов внутри последнего сообщения
        messages = driver.find_elements(By.CLASS_NAME, "qwen-chat-message-assistant")
        if messages:
            last_message = messages[-1]
            buttons = last_message.find_elements(By.TAG_NAME, "button")
            fallback_options = []
            for btn in buttons:
                if not btn.is_displayed() or not btn.is_enabled():
                    continue
                txt = btn.text.strip().lower()
                if any(kw in txt for kw in ["вариант", "option", "draft"]) or (txt.isdigit() and len(txt) == 1):
                    fallback_options.append(btn)
                    
            if len(fallback_options) >= 2:
                chosen_btn = random.choice(fallback_options[:2])
                chosen_text = chosen_btn.text.strip()
                logger.info(f"⚖️ Обнаружены альтернативные варианты ответа. Случайно выбираем: '{chosen_text}'")
                driver.execute_script("arguments[0].click();", chosen_btn)
                time.sleep(AFTER_CLICK_DELAY)
                return True
                
    except Exception as e:
        logger.debug(f"Не удалось обработать выбор из двух вариантов: {e}")
    return False

def get_prompt_template(excel_path):
    df = pd.read_excel(excel_path, sheet_name="Шаблон промта", header=None)
    if df.empty:
        raise ValueError("Лист 'Шаблон промта' пуст")
    template = df.iloc[0, 0]
    if pd.isna(template):
        raise ValueError("Ячейка A1 на листе 'Шаблон промта' пуста")
    return str(template)

def setup_webdriver():
    if not Path(EDGE_DRIVER_PATH).exists():
        logger.error(f"❌ EdgeDriver не найден по пути {EDGE_DRIVER_PATH}")
        return None
    options = Options()
    
    # Отключение логов C++ от Chromium-движка Edge в терминале
    options.add_argument("--log-level=3")
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    
    # Флаги оптимизации стабильности Edge при длительной работе
    options.add_argument("--disable-dev-shm-usage")  # Защита от переполнения общей памяти вкладок
    options.add_argument("--no-sandbox")             # Предотвращение системных сбоев на уровне контейнеров
    options.add_argument("--disable-gpu")            # Отключение аппаратного ускорения для снижения нагрузки на память
    
    # Предотвращение засыпания (Throttling) фоновых окон браузера в Windows
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    
    if HEADLESS_MODE:
        options.add_argument("--headless")
    options.add_argument(f"--window-size={WINDOW_WIDTH},{WINDOW_HEIGHT}")
    service = Service(executable_path=EDGE_DRIVER_PATH)
    driver = webdriver.Edge(service=service, options=options)
    return driver

def login_to_qwen(driver):
    try:
        logger.info("🔐 Запуск процесса авторизации в Qwen...")
        logger.debug("Ожидание кнопки 'Log in'...")
        login_button = WebDriverWait(driver, LOGIN_TIMEOUT).until(
            EC.element_to_be_clickable((By.XPATH, "//span[text()='Log in']/ancestor::button"))
        )
        login_button.click()
        logger.debug("Нажата кнопка 'Log in'")
        time.sleep(AFTER_CLICK_DELAY)

        email_input = WebDriverWait(driver, LOGIN_TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, "//input[@name='email' and @placeholder='Enter Your Email']"))
        )
        email_input.clear()
        time.sleep(AFTER_CLEAR_DELAY)
        # Оборонительное программирование: принудительно делаем строкой
        email_input.send_keys(str(QWEN_EMAIL))
        logger.debug("Email введён")

        password_input = driver.find_element(By.XPATH,
                                             "//input[@name='password' and @placeholder='Enter Your Password']")
        password_input.clear()
        time.sleep(AFTER_CLEAR_DELAY)
        password_input.send_keys(str(QWEN_PASSWORD))
        logger.debug("Пароль введён")
        time.sleep(1)

        sign_in_button = WebDriverWait(driver, LOGIN_TIMEOUT).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//span[text()='Sign in']/ancestor::button"))
        )
        sign_in_button.click()
        logger.debug("Нажата кнопка 'Sign in'")

        WebDriverWait(driver, LOGIN_TIMEOUT).until(
            EC.presence_of_element_located((By.CLASS_NAME, "message-input-textarea"))
        )
        logger.info("✅ Вход в систему Qwen успешно выполнен")
        return True
    except TimeoutException as e:
        logger.error(f"❌ Превышено время ожидания элементов формы при авторизации: {e}")
        return False
    except Exception as e:
        logger.error(f"❌ Непредвиденная ошибка в процессе авторизации: {e}")
        return False

def switch_to_fast_mode(driver):
    try:
        mode_selector = WebDriverWait(driver, SWITCH_MODE_TIMEOUT).until(
            EC.presence_of_element_located((By.CLASS_NAME, "qwen-select-thinking-label-text"))
        )
        current_mode = mode_selector.text.strip()
        logger.debug(f"Текущий режим модели: {current_mode}")
        if current_mode.lower() in ['быстрый', 'fast']:
            logger.debug("Режим Fast уже активен")
            return True
        mode_selector.click()
        time.sleep(AFTER_CLICK_DELAY)
        fast_mode_patterns = ['Быстрый', 'Fast', 'быстрый', 'fast']
        for pattern in fast_mode_patterns:
            try:
                fast_mode = driver.find_element(By.XPATH, f"//div[contains(@class, 'ant-select-item-option') and (contains(., '{pattern}') or contains(., '{pattern.lower()}'))]")
                fast_mode.click()
                logger.debug(f"Выбран режим: {pattern}")
                time.sleep(AFTER_CLICK_DELAY)
                return True
            except:
                continue
        options = driver.find_elements(By.CSS_SELECTOR, ".ant-select-item-option")
        if len(options) >= 2:
            options[1].click()
            logger.debug("Выбран второй режим (предположительно Fast/Быстрый)")
            time.sleep(AFTER_CLICK_DELAY)
            return True
        logger.warning("⚠️ Не удалось принудительно переключить режим модели")
        return False
    except Exception as e:
        logger.warning(f"⚠️ Не удалось обнаружить переключатель режимов модели: {e}")
        return False

def create_new_chat(driver):
    if not AUTO_CREATE_NEW_CHAT:
        return True
    try:
        logger.info("🔄 Запуск процесса создания нового чистого диалога...")
        new_chat_texts = ['Новый чат', 'New chat']
        for text in new_chat_texts:
            try:
                new_chat = WebDriverWait(driver, WAIT_FOR_NEW_CHAT_TIMEOUT).until(
                    EC.element_to_be_clickable((By.XPATH, f"//div[contains(@class, 'sidebar-entry-fixed-list-text') and text()='{text}']"))
                )
                new_chat.click()
                logger.info(f"🆕 Создан новый диалог (кнопка '{text}')")
                time.sleep(AFTER_CLICK_DELAY)
                switch_to_fast_mode(driver)
                WebDriverWait(driver, WAIT_FOR_ELEMENT_TIMEOUT).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "message-input-textarea"))
                )
                return True
            except:
                continue
        try:
            new_chat_container = WebDriverWait(driver, WAIT_FOR_NEW_CHAT_TIMEOUT).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "sidebar-entry-fixed-list-content"))
            )
            new_chat_container.click()
            logger.info("🆕 Создан новый диалог (клик по контейнеру иконки)")
            time.sleep(AFTER_CLICK_DELAY)
            switch_to_fast_mode(driver)
            WebDriverWait(driver, WAIT_FOR_ELEMENT_TIMEOUT).until(
                EC.presence_of_element_located((By.CLASS_NAME, "message-input-textarea"))
            )
            return True
        except:
            logger.warning("⚠️ Не удалось создать новый диалог через клик контейнера")
            return False
    except Exception as e:
        logger.error(f"❌ Ошибка в процессе создания нового диалога: {e}")
        return False

def wait_for_complete_response(driver, assistant_before, timeout=RESPONSE_TIMEOUT):
    WebDriverWait(driver, timeout).until(
        lambda d: len(d.find_elements(By.CLASS_NAME, "qwen-chat-message-assistant")) > assistant_before
    )
    
    start_time = time.time()
    while time.time() - start_time < timeout:
        messages = driver.find_elements(By.CLASS_NAME, "qwen-chat-message-assistant")
        if len(messages) > assistant_before:
            new_message = messages[-1]
            
            # 1. Немедленная проверка текста на системную ошибку прямо во время генерации (до появления Copy)
            current_text = new_message.text
            if has_error_message(current_text):
                logger.warning("⚠️ Обнаружен системный сбой или исчерпание лимита во время генерации. Прерываем ожидание.")
                return current_text
            
            # 2. Постоянно проверяем наличие двойного выбора на странице во время генерации
            if handle_dual_choices(driver):
                time.sleep(AFTER_CLICK_DELAY)
                continue
                
            # 3. Проверяем, завершилась ли генерация (появилась кнопка копирования)
            copy_elements = new_message.find_elements(By.CSS_SELECTOR, ".qwen-chat-package-comp-new-action-control-container-copy")
            if copy_elements:
                time.sleep(EXTRA_RESPONSE_DELAY)
                return new_message.text
                
        time.sleep(1)
        
    messages = driver.find_elements(By.CLASS_NAME, "qwen-chat-message-assistant")
    if len(messages) > assistant_before:
        return messages[-1].text
    raise TimeoutException("Таймаут ожидания ответа от модели")

def send_prompt_to_qwen(driver, prompt):
    textarea = WebDriverWait(driver, WAIT_FOR_ELEMENT_TIMEOUT).until(
        EC.presence_of_element_located((By.CLASS_NAME, "message-input-textarea"))
    )
    textarea.clear()
    time.sleep(AFTER_CLEAR_DELAY)
    
    # Безопасная работа с буфером обмена системы пользователя (сохраняем/восстанавливаем данные)
    user_clipboard = pyperclip.paste()
    try:
        pyperclip.copy(prompt)
        textarea.send_keys(Keys.CONTROL, 'v')
    finally:
        if user_clipboard:
            pyperclip.copy(user_clipboard)
            
    time.sleep(AFTER_PASTE_DELAY)
    assistant_before = len(driver.find_elements(By.CLASS_NAME, "qwen-chat-message-assistant"))
    textarea.send_keys(Keys.ENTER)
    time.sleep(AFTER_ENTER_DELAY)
    logger.info("Отправлено, ожидание ответа...")
    response = wait_for_complete_response(driver, assistant_before)
    return response

def send_with_timeout_handling(driver_container, prompt, class_name, batch_idx, batch_numbers):
    alternative_currencies = {
        '\u20b4': '₴', '\u20ba': '₺', '\u00a2': '¢', '\u00a4': '¤', '\u20a0': '₠'
    }
    for timeout_attempt in range(MAX_TIMEOUT_RETRIES + 1):
        try:
            # Перед любым действием проверяем живучесть браузера
            if not ensure_driver_active(driver_container):
                logger.error("❌ Не удалось восстановить сессию браузера. Пропускаем данный батч.")
                return None, False, batch_numbers
            
            driver = driver_container[0]
            response = send_prompt_to_qwen(driver, prompt)
            if not response:
                logger.warning("⚠️ Получен пустой ответ от модели.")
            else:
                response = response.replace('&euro;', '€').replace('&#8364;', '€')
                for alt_char in alternative_currencies.keys():
                    response = response.replace(alt_char, '€')

                # Немедленная обработка ошибок типа "что-то пошло не так" и исчерпания лимитов квоты
                if has_error_message(response):
                    logger.warning("⚠️ В чате обнаружена системная ошибка или исчерпание лимита квоты. Выполняется аварийный перезапуск диалога...")
                    if create_new_chat(driver):
                        continue
                    else:
                        return None, False, batch_numbers

                if CSV_SEPARATOR in response:
                    logger.info(f"🎯 Символ '{CSV_SEPARATOR}' успешно найден в ответе. Данные получены.")
                    return response, True, []
                else:
                    logger.warning(f"⚠️ В ответе отсутствует разделитель '{CSV_SEPARATOR}'. Начинаем попытки принудительного исправления...")
                    for retry in range(MAX_RETRIES):
                        retry_message = (
                            f"ВНИМАНИЕ! КРИТИЧЕСКАЯ ОШИБКА ФОРМАТА!\n"
                            f"Ты должен вернуть ТОЛЬКО CSV строки.\n"
                            f"РАЗДЕЛИТЕЛЬ СТРОГО '{CSV_SEPARATOR}'.\n"
                            f"УДАЛИ ЛЮБЫЕ слова вроде 'Вот ваш ответ', 'Готово' или извинения. "
                            f"ЕСЛИ НАПИШЕШЬ ХОТЬ ОДНО ЛИШНЕЕ СЛОВО — СИСТЕМА УПАДЕТ. ТОЛЬКО ДАННЫЕ!"
                        )
                        logger.info(f"🔄 Попытка исправления {retry + 1}/{MAX_RETRIES}: повторный запрос к Qwen...")
                        response = send_prompt_to_qwen(driver, retry_message)
                        if not response:
                            time.sleep(AFTER_CLICK_DELAY)
                            continue
                        
                        if has_error_message(response):
                            logger.warning("⚠️ Ошибка облака зафиксирована при исправлении. Аварийный сброс диалога...")
                            create_new_chat(driver)
                            break

                        response = response.replace('&euro;', '€').replace('&#8364;', '€')
                        for alt_char in alternative_currencies.keys():
                            response = response.replace(alt_char, '€')
                        if CSV_SEPARATOR in response:
                            logger.info(f"🎯 Повторный запрос успешен. Разделитель '{CSV_SEPARATOR}' получен.")
                            return response, True, []
                        time.sleep(AFTER_CLICK_DELAY)
                    return None, False, batch_numbers 

        except TimeoutException as e:
            logger.error(f"⌛ Таймаут {RESPONSE_TIMEOUT} сек. (попытка {timeout_attempt + 1} из {MAX_TIMEOUT_RETRIES + 1})")
            if timeout_attempt < MAX_TIMEOUT_RETRIES:
                driver = driver_container[0]
                logger.info("🔄 Перезапускаем чат для преодоления зависания...")
                if create_new_chat(driver):
                    continue
                else:
                    return None, False, batch_numbers
            else:
                return None, False, batch_numbers
        except Exception as e:
            logger.error(f"💥 Неожиданное исключение при отправке батча: {e}")
            # В случае вылета браузера или повреждения сессии форсируем очистку
            try:
                driver_container[0].quit()
            except:
                pass
            time.sleep(3)
            if timeout_attempt >= MAX_TIMEOUT_RETRIES:
                return None, False, batch_numbers
            
    return None, False, batch_numbers

def parse_response(response, expected_headers):
    """
    ОБОРОНИТЕЛЬНЫЙ ПАРСЕР: 
    Игнорирует любой текст от LLM и извлекает данные строго по их позиции.
    """
    data_lines = []
    
    # ШАГ 1: Извлекаем только строки с данными с помощью мощного Regex.
    # Ищем строки, которые начинаются с цифр (\d+), за которыми идет разделитель €.
    # Это мгновенно отсечет всю текстовую "болтовню" Qwen и заголовки таблицы!
    pattern = re.compile(rf'^(\d+)\s*{CSV_SEPARATOR}(.*)', re.MULTILINE)
    
    for match in pattern.finditer(response):
        num = match.group(1).strip()
        rest = match.group(2).strip()
        data_lines.append(f"{num}{CSV_SEPARATOR}{rest}")

    if not data_lines:
        logger.warning("⚠️ В ответе не найдено ни одной валидной строки данных (формат 'Номер€...').")
        return []

    # ШАГ 2: Собираем данные строго ПО ПОЗИЦИИ колонок (без привязки к именам заголовков)
    result = []
    for line in data_lines:
        parts = [p.strip() for p in line.split(CSV_SEPARATOR)]
        if len(parts) < 2:
            continue
            
        row = {}
        # Индекс 0 - всегда номер, Индекс 1 - всегда наименование
        row['номер'] = clean_id(parts[0])
        row['наименование'] = parts[1]
        
        # Индексы со 2-го и далее - это признаки из БД (строго по порядку)
        for i, exp_h in enumerate(expected_headers):
            idx = i + 2
            row[exp_h.lower()] = parts[idx] if idx < len(parts) else ''
            
        # Самый последний элемент (если он есть) - это всегда Уверенность
        if len(parts) > 2:
            row['уверенность'] = parts[-1]
        else:
            row['уверенность'] = ''
            
        result.append(row)
        
    return result

def save_debug_file(class_name, batch_idx, filename_suffix, content):
    if not SAVE_DEBUG_FILES: return
    debug_dir = Path(DEBUG_DIR) / class_name
    debug_dir.mkdir(parents=True, exist_ok=True)
    filepath = debug_dir / f"batch_{batch_idx+1}_{filename_suffix}.txt"
    with open(filepath, "w", encoding=EXPECTED_ENCODING) as f:
        f.write(content)

def log_response_to_class_file(class_name, batch_idx, response):
    if not SAVE_DEBUG_FILES: return
    debug_dir = Path(DEBUG_DIR) / class_name
    debug_dir.mkdir(parents=True, exist_ok=True)
    log_file = debug_dir / "responses.log"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a", encoding=EXPECTED_ENCODING) as f:
        if class_name not in _session_logged_for_class:
            f.write(f"\n{'#'*70}\n## НАЧАЛО СЕАНСА: {SESSION_START}\n{'#'*70}\n")
            _session_logged_for_class.add(class_name)
        f.write(f"\n{'='*70}\nБатч {batch_idx+1}  |  {timestamp}\n{'='*70}\n{response}\n")

def post_process_excel(file_path):
    if not Path(file_path).exists(): return
    wb = load_workbook(file_path)
    ws = wb.active

    if ws is None:
        logger.error(f"❌ Ошибка: В файле {file_path} не найден активный лист!")
        return

    rows_data = [list(row) for row in ws.iter_rows(values_only=True)]
    if not rows_data: return

    header_indices = [idx for idx, row in enumerate(rows_data) if row and row[0] == 'ид_исторический']
    if len(header_indices) <= 1: return

    blocks = []
    for i, h_idx in enumerate(header_indices):
        start_data = h_idx + 1
        end_data = header_indices[i+1] if i+1 < len(header_indices) else len(rows_data)
        blocks.append((h_idx, list(range(start_data, end_data))))

    class_data = {}
    for h_idx, data_idxs in blocks:
        if not data_idxs: continue
        header_row = rows_data[h_idx]
        try:
            class_col_idx = header_row.index('класс')
        except ValueError: continue

        first_data_row = rows_data[data_idxs[0]]
        if len(first_data_row) <= class_col_idx: continue
        class_name = first_data_row[class_col_idx] or ""

        if class_name not in class_data:
            class_data[class_name] = {'header': header_row, 'rows': []}
        class_data[class_name]['rows'].extend([rows_data[i] for i in data_idxs])

    def sort_key(value):
        if value is None: return (2, '')
        try: return (0, int(value))
        except (ValueError, TypeError): return (1, str(value))

    for class_name, data in class_data.items():
        header = data['header']
        try:
            id_col_idx = header.index('ид_исторический')
            data['rows'].sort(key=lambda r: sort_key(r[id_col_idx]))
        except ValueError: continue

    new_rows = []
    for cls in sorted(class_data.keys()):
        new_rows.append(class_data[cls]['header'])
        new_rows.extend(class_data[cls]['rows'])

    ws.delete_rows(1, ws.max_row)
    for r_idx, row_vals in enumerate(new_rows, start=1):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
            
    # Применение безопасного сохранения после пост-обработки
    safe_save_workbook(wb, file_path)

def filter_materials_by_brand(excel_path, output_missing_path):
    df_materials = pd.read_excel(excel_path, sheet_name="Материалы")
    df_brands = pd.read_excel(excel_path, sheet_name="Справочник брендов")

    df_materials.rename(columns=lambda x: str(x).strip(), inplace=True)
    df_brands.rename(columns=lambda x: str(x).strip(), inplace=True)

    mat_cols = {col.lower(): col for col in df_materials.columns}
    if 'бренд' in mat_cols:
        df_materials.rename(columns={mat_cols['бренд']: 'Бренд'}, inplace=True)
    else:
        logger.error("В листе 'Материалы' отсутствует колонка 'Бренд'.")
        return pd.DataFrame(), pd.DataFrame()

    brand_cols = {col.lower(): col for col in df_brands.columns}
    if 'бренд' in brand_cols and 'сайт' in brand_cols:
        df_brands.rename(columns={brand_cols['бренд']: 'Бренд', brand_cols['сайт']: 'Сайт'}, inplace=True)
    else:
        logger.error("В листе 'Справочник брендов' должны быть колонки 'Бренд' и 'Сайт'.")
        return pd.DataFrame(), pd.DataFrame()

    valid_brands = set(df_brands['Бренд'].dropna().unique())
    mask = df_materials['Бренд'].isin(valid_brands)
    valid_df = df_materials[mask].copy()
    invalid_df = df_materials[~mask].copy()

    if not invalid_df.empty:
        # Для второстепенного файла отбракованных брендов также добавим базовый перехват
        try:
            invalid_df.to_excel(output_missing_path, index=False)
        except PermissionError:
            logger.warning(f"⚠️ Не удалось сохранить {output_missing_path}, так как он заблокирован.")
        logger.warning(f"Отбраковано {len(invalid_df)} записей с брендами, отсутствующими в справочнике.")
    return valid_df, df_brands

def create_database_from_filtered_materials(valid_materials_df, df_brands, db_path):
    if Path(db_path).exists():
        Path(db_path).unlink()

    df_materials = valid_materials_df.copy()
    rename_map = {}
    for col in df_materials.columns:
        col_lower = str(col).strip().lower()
        if col_lower == 'наименование': rename_map[col] = 'наименование'
        elif col_lower in ['ид.исторический', 'ид_исторический']: rename_map[col] = 'ид_исторический'
        elif col_lower == 'класс': rename_map[col] = 'класс'
        elif col_lower == 'номер': rename_map[col] = 'номер'
        elif col_lower == 'бренд': rename_map[col] = 'бренд'
    
    if rename_map: df_materials.rename(columns=rename_map, inplace=True)

    if 'номер' not in df_materials.columns:
        df_materials.insert(0, "номер", range(1, len(df_materials) + 1))

    df_brands_clean = df_brands[['Бренд', 'Сайт']].copy()
    df_brands_clean.columns = ['бренд', 'сайт']

    df_class_priznaki = pd.read_excel(INPUT_EXCEL, sheet_name="Класс-Признаки")
    col_class, col_priznak = df_class_priznaki.columns[0], df_class_priznaki.columns[1]
    df_class_priznaki = df_class_priznaki[[col_class, col_priznak]].dropna()
    df_class_priznaki.columns = ["class_name", "priznak"]
    all_priznaki = sorted(df_class_priznaki["priznak"].unique())

    conn = sqlite3.connect(db_path)
    df_materials.to_sql("materials", conn, if_exists="replace", index=False)
    df_brands_clean.to_sql("brands", conn, if_exists="replace", index=False)
    df_class_priznaki.to_sql("class_priznaki", conn, if_exists="replace", index=False)

    columns_def = [
        "номер INTEGER", "наименование TEXT", "ид_исторический TEXT",
        "класс TEXT", "бренд TEXT", "статус TEXT DEFAULT 'OK'", "уверенность TEXT"
    ]
    
    for p in all_priznaki: 
        safe_p = str(p).replace('"', '""')
        columns_def.append(f'"{safe_p}" TEXT')
    
    conn.execute(f"CREATE TABLE results ({', '.join(columns_def)})")
    conn.execute("CREATE TABLE IF NOT EXISTS processed (номер INTEGER PRIMARY KEY, class_name TEXT, brand TEXT, processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    conn.execute("CREATE TABLE IF NOT EXISTS retry_counter (номер INTEGER PRIMARY KEY, class_name TEXT, brand TEXT, attempts INTEGER DEFAULT 0)")
    
    # Настройка быстрого режима записи WAL для SQLite (предотвращает зависания при транзакциях)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.close()

def get_all_numbers_from_materials(db_path):
    with sqlite3.connect(db_path) as conn:
        return {row[0] for row in conn.execute("SELECT номер FROM materials").fetchall()}

def get_processed_numbers(db_path):
    with sqlite3.connect(db_path) as conn:
        return {row[0] for row in conn.execute("SELECT номер FROM processed").fetchall()}

def mark_processed(db_path, numbers, class_name, brand):
    if not numbers: return
    with sqlite3.connect(db_path) as conn:
        conn.executemany("INSERT OR IGNORE INTO processed (номер, class_name, brand) VALUES (?, ?, ?)", [(num, class_name, brand) for num in numbers])

def save_results_to_db(db_path, rows_dict_list):
    if not rows_dict_list: return
    with sqlite3.connect(db_path) as conn:
        existing_columns = [row[1] for row in conn.execute("PRAGMA table_info(results)").fetchall()]
        filtered_rows = [{k: v for k, v in row.items() if k in existing_columns} for row in rows_dict_list]
        pd.DataFrame(filtered_rows).to_sql("results", conn, if_exists="append", index=False)

def update_status_for_numbers(db_path, numbers, status):
    if not numbers: return
    with sqlite3.connect(db_path) as conn:
        conn.execute(f"UPDATE results SET статус = ? WHERE номер IN ({','.join('?' for _ in numbers)})", [status] + list(numbers))

def insert_missing_rows_as_error(db_path, missing_numbers, class_names_map, brands_map):
    if not missing_numbers: return
    with sqlite3.connect(db_path) as conn:
        all_priznaki = [row[0] for row in conn.execute("SELECT DISTINCT priznak FROM class_priznaki").fetchall()]
        rows_to_insert = []
        for num in missing_numbers:
            row = conn.execute("SELECT наименование, ид_исторический FROM materials WHERE номер = ?", (num,)).fetchone()
            base_row = {
                'номер': num, 'наименование': row[0] if row else '', 'ид_исторический': row[1] if row else '',
                'класс': class_names_map.get(num, ''), 'бренд': brands_map.get(num, ''), 'статус': 'ошибка', 'уверенность': ''
            }
            for p in all_priznaki: base_row[p] = ''
            rows_to_insert.append(base_row)
        if rows_to_insert:
            pd.DataFrame(rows_to_insert).to_sql("results", conn, if_exists="append", index=False)

def append_class_brand_block_to_excel(output_file, class_name, brand, data_rows, priznaki):
    columns_order = ['ид_исторический', 'наименование', 'класс', 'бренд', 'статус', 'уверенность'] + priznaki
    df_block = pd.DataFrame(data_rows).reindex(columns=columns_order, fill_value='')

    if Path(output_file).exists():
        book = load_workbook(output_file)
        sheet = book.active
        
        # === ДОБАВЛЕНА БРОНЯ ===
        if sheet is None:
            logger.error("Лист не найден, создаем новый")
            sheet = book.create_sheet("Sheet")
        # =======================
            
        start_row = sheet.max_row + 1
    else:
        book, start_row = Workbook(), 1
        sheet = book.active
        # === ДОБАВЛЕНА БРОНЯ ===
        if sheet is None:
            sheet = book.create_sheet("Sheet")

        
    for col_idx, col_name in enumerate(columns_order, start=1):
        sheet.cell(row=start_row, column=col_idx, value=col_name)

    for row_idx, row in enumerate(df_block.itertuples(index=False), start=start_row+1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
            
    # Безопасное сохранение вместо book.save()
    safe_save_workbook(book, output_file)

def get_class_brand_groups(db_path):
    with sqlite3.connect(db_path) as conn:
        return list(pd.read_sql_query("SELECT DISTINCT класс, бренд FROM materials ORDER BY класс, бренд", conn).itertuples(index=False, name=None))

def get_data_by_class_brand(db_path, class_name, brand, exclude_processed=True):
    processed = get_processed_numbers(db_path) if exclude_processed else set()
    with sqlite3.connect(db_path) as conn:
        df = pd.read_sql_query("SELECT номер, наименование, ид_исторический FROM materials WHERE класс = ? AND бренд = ? ORDER BY номер", conn, params=(class_name, brand))
    return df[~df['номер'].isin(processed)] if exclude_processed and processed else df

def get_brand_site(db_path, brand):
    with sqlite3.connect(db_path) as conn:
        row = conn.execute("SELECT сайт FROM brands WHERE бренд = ?", (brand,)).fetchone()
    return row[0] if row else ''

def get_priznaki_for_class(db_path, class_name):
    with sqlite3.connect(db_path) as conn:
        return [row[0] for row in conn.execute("SELECT priznak FROM class_priznaki WHERE class_name = ? ORDER BY priznak", (class_name,)).fetchall()]

def get_all_priznaki_from_db(db_path):
    with sqlite3.connect(db_path) as conn:
        return [row[0] for row in conn.execute("SELECT DISTINCT priznak FROM class_priznaki ORDER BY priznak").fetchall()]

def get_class_name_for_number(db_path, number):
    with sqlite3.connect(db_path) as conn:
        row = conn.execute("SELECT класс, бренд FROM materials WHERE номер = ?", (number,)).fetchone()
    return row if row else (None, None)

# ==================== ОСНОВНАЯ ФУНКЦИЯ ====================
def main():
    if not Path(INPUT_EXCEL).exists():
        logger.error(f"❌ Файл {INPUT_EXCEL} не найден")
        return

    valid_materials_df, brands_df = filter_materials_by_brand(INPUT_EXCEL, OUTPUT_MISSING_BRAND)
    if valid_materials_df.empty:
        return

    try:
        create_database_from_filtered_materials(valid_materials_df, brands_df, DB_PATH)
        # Проверка и импорт уже выполненных записей из выходного файла перед стартом
        load_already_processed_from_excel(OUTPUT_EXCEL, DB_PATH)
        template = get_prompt_template(INPUT_EXCEL)
    except Exception as e:
        logger.error(f"❌ Ошибка инициализации БД или шаблона: {e}")
        return

    total_items = len(get_all_numbers_from_materials(DB_PATH))
    accounted_numbers = set(get_processed_numbers(DB_PATH))
    
    # Если все материалы из входного файла уже есть в выходном, завершаем работу
    if len(accounted_numbers) >= total_items:
        logger.info("🎉 Все материалы уже успешно обработаны и находятся в выходном файле. Завершение работы.")
        return

    driver = setup_webdriver()
    if driver is None: return

    # Обернем driver в контейнер для сквозного отслеживания и автовосстановления внутри функций
    driver_container = [driver]

    try:
        driver.get("https://chat.qwen.ai")
        time.sleep(PAGE_LOAD_DELAY)

        if not login_to_qwen(driver):
            driver.quit()
            return

        WebDriverWait(driver, WAIT_FOR_ELEMENT_TIMEOUT).until(EC.presence_of_element_located((By.CLASS_NAME, "message-input-textarea")))
        time.sleep(AFTER_CLICK_DELAY)
        switch_to_fast_mode(driver)

        class_brand_groups = get_class_brand_groups(DB_PATH)
        pbar = tqdm(total=total_items, initial=len(accounted_numbers), desc="✅ Общий прогресс", unit="номер", position=0)

        global_batch_counter = 0

        for class_name, brand in class_brand_groups:
            site = get_brand_site(DB_PATH, brand)
            priznaki = get_priznaki_for_class(DB_PATH, class_name)
            if not site or not priznaki: continue

            df_group_data = get_data_by_class_brand(DB_PATH, class_name, brand)
            if df_group_data.empty: continue

            for batch_idx in range((len(df_group_data) + BATCH_SIZE - 1) // BATCH_SIZE):
                batch_df = df_group_data.iloc[batch_idx * BATCH_SIZE:(batch_idx + 1) * BATCH_SIZE]
                batch_numbers = batch_df['номер'].tolist()
                items_text = "\n".join([f"{row['номер']}. {row['наименование']}" for _, row in batch_df.iterrows()])

                prompt = template.replace("<А1 - взять по классу Класс-признаки>", ", ".join(priznaki)) \
                                 .replace("<А3 - взять по бренду Справочник брендов>", site) \
                                 .replace("<А2 - взять 500 материалов или меньше по классу + бренду>", items_text)

                # Передаем контейнер с браузером вместо обычного объекта
                response, success, _ = send_with_timeout_handling(driver_container, prompt, class_name, batch_idx, batch_numbers)
                if not success: continue

                parsed_rows = parse_response(response, priznaki)
                if parsed_rows:
                    good_numbers = [str(n) for n in batch_numbers if str(n) in {str(r.get('номер', '')).strip() for r in parsed_rows if r.get('номер')}]
                    if good_numbers:
                        db_rows, excel_rows = [], []
                        id_map = {str(k): v for k, v in zip(batch_df['номер'], batch_df['ид_исторический'])}
                        
                        for row in parsed_rows:
                            if row.get('номер') not in good_numbers: continue
                            db_row = {'номер': int(row['номер']), 'наименование': row.get('наименование', ''), 'ид_исторический': id_map.get(row['номер'], ''), 'класс': class_name, 'бренд': brand, 'статус': 'OK', 'уверенность': row.get('уверенность', '')}
                            excel_row = {'ид_исторический': id_map.get(row['номер'], ''), 'наименование': row.get('наименование', ''), 'класс': class_name, 'бренд': brand, 'статус': 'OK', 'уверенность': row.get('уверенность', '')}
                            
                            for p in get_all_priznaki_from_db(DB_PATH): db_row[p] = row.get(p.lower(), '')
                            for p in priznaki: excel_row[p] = row.get(p.lower(), '')
                            
                            db_rows.append(db_row)
                            excel_rows.append(excel_row)

                        if db_rows:
                            save_results_to_db(DB_PATH, db_rows)
                            mark_processed(DB_PATH, [int(n) for n in good_numbers], class_name, brand)
                            append_class_brand_block_to_excel(OUTPUT_EXCEL, class_name, brand, excel_rows, priznaki)
                            new_nums = set(int(n) for n in good_numbers) - accounted_numbers
                            accounted_numbers.update(new_nums)
                            pbar.update(len(new_nums))

                global_batch_counter += 1
                if AUTO_CREATE_NEW_CHAT and global_batch_counter >= NEW_CHAT_EVERY_N_BATCHES:
                    # Извлекаем текущий (возможно обновленный) объект драйвера из контейнера
                    current_driver = driver_container[0]
                    create_new_chat(current_driver)
                    global_batch_counter = 0
                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

        post_process_excel(OUTPUT_EXCEL)
        pbar.close()

    except Exception as e:
        logger.error(f"❌ Критическая ошибка выполнения программы: {e}", exc_info=True)
    finally:
        try:
            driver_container[0].quit()
        except:
            pass

if __name__ == "__main__":
    main()